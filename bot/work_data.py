import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
wb = load_workbook(filename=f'{BASE_DIR}/way_to_dream.xlsx')
ws = wb['новая сводка']

data = json.load(open(f'{BASE_DIR}/bot/count_run.json'))
n = data["count_run"]

def type(type) -> str:  # вводим на чем заработали
    global n

    data = json.load(open(f'{BASE_DIR}/bot/count_run.json'))
    n = data["count_run"]

    type_link = f"G{2+n}"  # на чем заработал

    ws[type_link] = type  # сохраним в ячейку с датой сегодняшнюю
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')


def dates():  # сохраним дату сегодня
    dates_link = f"A{2+n}"  # ссылка на ячейку с датой
    today = date.today()
    ws[dates_link] = today  # сохраним в ячейку с датой сегодняшнюю
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')


def money(money) -> int:  # запишем сколько заработали
    money_link = f"B{2+n}"  # ссылка на сколько заработал

    ws[money_link] = money  # сохраним в ячейку с датой сегодняшнюю
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')


def counts(count) -> int:  # просим ввести сколько заказов
    counts_link = f"C{2+n}"  # сколько заказов

    ws[counts_link] = count  # сохраним в ячейку с датой сегодняшнюю
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')


def start():  # записываем время сейчас старт
    start_link = f"D{3+n}"  # колхоз начало работы

    now = datetime.now()
    ws[start_link] = now.strftime("%H:%M")
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')


def stop():  # записываем время конца
    global n

    stop_link = f"D{4+n}"  # колхоз конец  работы
    time_link = f"D{2+n}"  # разница времени работы
    start_link = f"D{3+n}"  # колхоз начало работы

    now = datetime.now()
    ws[stop_link] = now.strftime("%H:%M")
    ws[time_link] = f'={stop_link}-{start_link}'
    wb.save(f'{BASE_DIR}/way_to_dream.xlsx')

    n += 1
    data = {"count_run": n}
    with open(f'{BASE_DIR}/bot/count_run.json', 'w') as f:
        json.dump(data, f)
